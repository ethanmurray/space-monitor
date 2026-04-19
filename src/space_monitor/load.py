"""Per-sheet loaders that turn ``Space_Dashboard_Hardcopy.xlsx`` into rows in
the SQLite schema defined in :mod:`space_monitor.db`.

The loader runs against a freshly-initialized database (see :func:`load_all`).
Each ``_load_<sheet>`` function is independent; if a sheet fails, it raises and
no partial state leaks because the caller wraps everything in a transaction.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Iterable

import openpyxl
import warnings

from . import db, taxonomy

warnings.filterwarnings("ignore", message="Sparkline Group extension is not supported")
warnings.filterwarnings("ignore", message="Data Validation extension is not supported")


# ---------------------------------------------------------------------------
# Cell coercion helpers
# ---------------------------------------------------------------------------


def _txt(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.lower() in {"n/a", "na"}:
        return None
    return s


def _int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _flt(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _bool_flag(v) -> int | None:
    """Yes/No/blank -> 1/0/None. Some legend cells carry literal 1/0 already."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in {"yes", "y", "true", "1"}:
        return 1
    if s in {"no", "n", "false", "0"}:
        return 0
    return None


# ---------------------------------------------------------------------------
# Reference / taxonomy load (from JSON)
# ---------------------------------------------------------------------------


def _load_taxonomy(conn: sqlite3.Connection) -> None:
    t = taxonomy.load()
    cur = conn.cursor()
    cur.executemany("INSERT INTO cocom(name) VALUES (?)", [(c,) for c in t.cocoms])
    cur.executemany(
        "INSERT INTO country(name, cocom, priority) VALUES (?, ?, ?)",
        [(c.name, c.cocom, 1 if c.priority else 0) for c in t.countries],
    )
    cur.executemany(
        "INSERT INTO partnership_type(name, score) VALUES (?, ?)",
        [(t_.name, int(t_.score)) for t_ in t.partnership_types],
    )
    cur.executemany(
        "INSERT INTO level_of_commitment(name, score) VALUES (?, ?)",
        [(t_.name, int(t_.score)) for t_ in t.levels_of_commitment],
    )
    cur.executemany(
        "INSERT INTO business_model(name, score) VALUES (?, ?)",
        [(t_.name, int(t_.score)) for t_ in t.business_models],
    )
    cur.executemany(
        "INSERT INTO mission_type(name, score) VALUES (?, ?)",
        [(t_.name, int(t_.score)) for t_ in t.mission_types],
    )
    cur.executemany(
        "INSERT INTO relationship_type(name) VALUES (?)",
        [(n,) for n in t.relationship_types],
    )
    cur.executemany(
        "INSERT INTO operator_type(name) VALUES (?)",
        [(n,) for n in t.operator_types],
    )
    cur.executemany(
        "INSERT INTO organization_type(name) VALUES (?)",
        [(n,) for n in t.organization_types],
    )
    cur.executemany(
        "INSERT INTO sovereign_status(name) VALUES (?)",
        [(n,) for n in t.sovereign_statuses],
    )
    cur.executemany(
        "INSERT INTO mass_class(name, lower_kg, upper_kg) VALUES (?, ?, ?)",
        [(m.name, m.lower_kg, m.upper_kg) for m in t.mass_classes],
    )
    cur.executemany("INSERT INTO orbit(name) VALUES (?)", [(o,) for o in t.orbits])
    cur.executemany(
        "INSERT INTO mission_area(name) VALUES (?)", [(m,) for m in t.mission_areas]
    )
    cur.executemany(
        "INSERT INTO value_chain_segment(name) VALUES (?)",
        [(v,) for v in t.value_chain_segments],
    )
    cur.executemany(
        """
        INSERT INTO partnership_strength_lookup
            (partnership_type, business_model, mission_type, score)
        VALUES (?, ?, ?, ?)
        """,
        [
            (s.partnership_type, s.business_model, s.mission_type, s.score)
            for s in t.partnership_strength_lookup
        ],
    )


# ---------------------------------------------------------------------------
# Domain loaders (one per sheet)
# ---------------------------------------------------------------------------


def _rows(ws, start: int) -> Iterable[tuple]:
    """Iterate value rows from `start` (1-indexed) onward."""
    return ws.iter_rows(min_row=start, values_only=True)


def _load_universal_legend(conn, ws) -> int:
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    segments = [h for h in headers[2:] if h]  # skip Country, COCOM
    rows = []
    for r in _rows(ws, 2):
        country = _txt(r[0])
        if not country:
            continue
        for idx, seg in enumerate(segments, start=2):
            flag = _bool_flag(r[idx]) if idx < len(r) else None
            if flag is not None:
                rows.append((country, seg, flag))
    conn.executemany(
        "INSERT OR REPLACE INTO country_capability(country, segment, applicable) VALUES (?, ?, ?)",
        rows,
    )
    return len(rows)


def _load_budget_toplines(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 2):
        country = _txt(r[0])
        cd = _txt(r[1])
        if not country or not cd:
            continue
        rows.append(
            (
                country,
                cd,
                _txt(r[2]),
                _txt(r[3]),
                _flt(r[4]),
                _txt(r[5]),
                _flt(r[6]),
            )
        )
    conn.executemany(
        """
        INSERT OR REPLACE INTO budget_topline
            (country, civil_or_defense, estimate_or_exact, method,
             topline_spend_usd, cagr_range, cagr)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_defense_spending(conn, ws) -> int:
    headers = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    # year columns start at index 6
    year_cols = [(idx, _int(h)) for idx, h in enumerate(headers) if _int(h) and _int(h) > 1900]
    rows = []
    for r in _rows(ws, 3):
        country = _txt(r[2])
        if not country:
            continue
        region = _txt(r[1])
        line_type = _txt(r[3])
        funding_source = _txt(r[4])
        account_type = _txt(r[5])
        for idx, year in year_cols:
            amt = _flt(r[idx]) if idx < len(r) else None
            if amt is None:
                continue
            rows.append(
                (region, country, line_type, funding_source, account_type, year, amt)
            )
    conn.executemany(
        """
        INSERT OR REPLACE INTO defense_spending
            (region, country, line_type, funding_source, account_type, year, amount_usd_thousands)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_space_program_investments(conn, ws) -> int:
    # Year column header lives in row 3 alongside the rest of the field names;
    # row 2 is a banner whose merged cells leave the year labels misaligned by
    # one column when read non-rendered, so reading row 3 is the safe path.
    header_year_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    year_cols = []
    for idx, h in enumerate(header_year_row):
        y = _int(h)
        if y and y > 1900:
            year_cols.append((idx, y))

    program_rows = []
    year_rows: list[tuple[int, int, float]] = []
    cur = conn.cursor()
    for r in _rows(ws, 4):
        country = _txt(r[1])
        program = _txt(r[4])
        if not country and not program:
            continue
        cur.execute(
            """
            INSERT INTO space_program_investment
                (country, organization, partnership_id, program, technologies,
                 primary_mission, secondary_mission, overview,
                 link_1, link_2, link_3,
                 total_funding_local_m, local_currency, total_funding_usd_m,
                 start_year, end_year, total_years)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                country,
                _txt(r[2]),
                _txt(r[3]),
                program,
                _txt(r[6]),
                _txt(r[7]),
                _txt(r[8]),
                _txt(r[9]),
                _txt(r[10]),
                _txt(r[11]),
                _txt(r[12]),
                _flt(r[13]),
                _txt(r[14]),
                _flt(r[15]),
                _int(r[16]),
                _int(r[17]),
                _int(r[18]),
            ),
        )
        program_id = cur.lastrowid
        program_rows.append(program_id)
        for idx, year in year_cols:
            amt = _flt(r[idx]) if idx < len(r) else None
            if amt is None:
                continue
            year_rows.append((program_id, year, amt))

    cur.executemany(
        """
        INSERT OR REPLACE INTO space_program_investment_year(program_id, year, amount)
        VALUES (?, ?, ?)
        """,
        year_rows,
    )
    return len(program_rows)


def _load_industrial_base_scores(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 3):
        country = _txt(r[1])
        seg = _txt(r[2])
        if not country or not seg:
            continue
        rows.append(
            (
                country,
                seg,
                _int(r[3]),
                _int(r[4]),
                _int(r[5]),
                _int(r[6]),
                _int(r[7]),
                _int(r[8]),
                _int(r[9]),
                _txt(r[10]),
            )
        )
    conn.executemany(
        """
        INSERT OR REPLACE INTO industrial_base_score
            (country, value_chain_segment, mission_agnostic_launch_only,
             comms_data_transport, isr_remote_sensing, missile_warning_tracking,
             space_domain_awareness, science_exploration, pnt, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_space_assets(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 4):
        # Skip blank rows.
        if not any(r[2:6]):
            continue
        rows.append(
            (
                _txt(r[2]),  # profile_country
                _txt(r[3]),  # asset_name
                _txt(r[4]),  # spacecraft_name
                _txt(r[5]),  # owner
                _txt(r[6]),  # operator
                _txt(r[7]),  # operator_type
                _txt(r[8]),  # mission_type
                _txt(r[9]),  # sovereign_status
                _txt(r[10]),  # prime_contractor_country
                _txt(r[11]),  # prime_contractor
                _txt(r[12]),  # subcontractor_1
                _txt(r[13]),  # subcontractor_2
                _txt(r[14]),  # launch_provider
                _txt(r[15]),  # launch_vehicle_class
                _txt(r[16]),  # launch_vehicle
                _int(r[17]),  # num_current
                _int(r[18]),  # num_planned
                _int(r[19]),  # num_total
                _int(r[20]),  # order_year
                _int(r[21]),  # ioc
                _int(r[22]),  # foc
                _int(r[23]),  # launch_year
                _int(r[24]),  # actual_eol
                _int(r[25]),  # assumed_eol
                _int(r[26]),  # eol_year
                _int(r[27]),  # constellation_launch_year
                _int(r[28]),  # constellation_assets
                _flt(r[29]),  # coverage_score
                _flt(r[30]),  # mass_score
                _flt(r[31]),  # launch_year_score
                _flt(r[32]),  # capability_score
                _flt(r[33]),  # total_score
                _flt(r[34]),  # final_score
                _bool_flag(r[36]),  # SATCOM
                _bool_flag(r[37]),  # PNT
                _bool_flag(r[38]),  # ISR
                _bool_flag(r[39]),  # Environmental
                _bool_flag(r[40]),  # Missile warning
                _bool_flag(r[41]),  # SDA
                _bool_flag(r[42]),  # Combat power
                _bool_flag(r[43]),  # Orbital
                _bool_flag(r[44]),  # Lunar
                _bool_flag(r[45]),  # Interplanetary
                _bool_flag(r[46]),  # BMC3
                _bool_flag(r[47]),  # Launch
                _bool_flag(r[48]),  # OOS
                _bool_flag(r[49]),  # Other
                _bool_flag(r[50]),  # Not specified
                _txt(r[51]),  # orbit
                _flt(r[52]),  # mass
                _txt(r[53]),  # mass_class
                _txt(r[54]),  # description
                _txt(r[55]),  # link_1
                _txt(r[56]),  # link_2
                _txt(r[57]),  # link_3
                _txt(r[58]),  # link_4
            )
        )
    conn.executemany(
        f"""
        INSERT INTO space_asset (
            profile_country, asset_name, spacecraft_name, owner, operator,
            operator_type, mission_type, sovereign_status,
            prime_contractor_country, prime_contractor,
            subcontractor_1, subcontractor_2,
            launch_provider, launch_vehicle_class, launch_vehicle,
            num_current_assets, num_planned_assets, num_total_assets,
            order_year, ioc, foc, launch_year,
            actual_eol, assumed_eol, eol_year,
            constellation_launch_year, constellation_assets,
            coverage_score, mass_score, launch_year_score,
            capability_score, total_score, final_score,
            mission_satcom, mission_pnt, mission_isr, mission_environmental,
            mission_missile_warning, mission_sda, mission_combat_power,
            mission_orbital, mission_lunar, mission_interplanetary,
            mission_bmc3, mission_launch, mission_oos, mission_other,
            mission_not_specified,
            orbit, mass_kg, mass_class, description,
            link_1, link_2, link_3, link_4
        ) VALUES ({", ".join(["?"] * 56)})
        """,
        rows,
    )
    return len(rows)


def _load_partnerships(conn, ws) -> int:
    rows = []
    seen: set[str] = set()
    for r in _rows(ws, 4):
        pid = _txt(r[3])
        if not pid or pid in seen:
            continue
        seen.add(pid)
        rows.append(
            (
                pid,
                _txt(r[5]),  # parent_id
                _txt(r[1]),  # analyst
                _txt(r[2]),  # new/legacy
                _txt(r[4]),  # description
                _txt(r[9]),  # primary_mission
                _txt(r[10]),  # sub_mission
                _int(r[11]),  # year
                _txt(r[12]),  # partnership_type
                _txt(r[13]),  # level_of_commitment
                _txt(r[14]),  # relationship_type
                _txt(r[15]),  # business_model
                _txt(r[16]),  # mission_type
                _flt(r[17]),  # strength
                _txt(r[18]),  # asset_name
                _txt(r[19]),  # cocom_1
                _txt(r[20]),  # country_1
                _txt(r[21]),  # org_type_1
                _txt(r[22]),  # organization_1
                _txt(r[23]),  # company_1
                _txt(r[24]),  # cocom_2
                _txt(r[25]),  # country_2
                _txt(r[26]),  # org_type_2
                _txt(r[27]),  # organization_2
                _txt(r[28]),  # company_2
                _txt(r[29]),  # link_1
                _txt(r[30]),  # link_2
                _txt(r[31]),  # link_3
            )
        )
    conn.executemany(
        f"""
        INSERT OR REPLACE INTO partnership (
            partnership_id, parent_id, analyst, new_or_legacy, description,
            primary_mission, sub_mission, partnership_year,
            partnership_type, level_of_commitment, relationship_type,
            business_model, mission_type, partnership_strength, asset_name,
            cocom_1, country_1, org_type_1, organization_1, company_1,
            cocom_2, country_2, org_type_2, organization_2, company_2,
            link_1, link_2, link_3
        ) VALUES ({", ".join(["?"] * 28)})
        """,
        rows,
    )
    return len(rows)


def _load_industry(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 4):
        bu = _txt(r[4])
        country = _txt(r[3])
        if not bu and not country:
            continue
        rows.append(
            (
                _txt(r[2]),  # priority_country
                country,
                bu,
                _txt(r[5]),  # bu_hq_city
                _txt(r[6]),  # bu_hq_country
                _int(r[7]),  # founding_year
                _txt(r[8]),  # parent
                _txt(r[9]),  # parent_city
                _txt(r[10]),  # parent_country
                _txt(r[11]),  # primary_value_chain
                _txt(r[12]),  # cap_satellite_integration
                _txt(r[13]),  # cap_spacecraft_components
                _txt(r[14]),  # cap_payload_sensors
                _txt(r[15]),  # cap_digital_services
                _txt(r[16]),  # cap_launch
                _txt(r[17]),  # cap_ground
                _txt(r[18]),  # cap_other
                _txt(r[19]),  # primary_mission
                _txt(r[20]),  # sub_mission
                _txt(r[21]),  # mission_satcom_pnt
                _txt(r[22]),  # mission_space_sensing
                _txt(r[23]),  # mission_sda_combat_power
                _txt(r[24]),  # mission_science_exploration
                _txt(r[25]),  # mission_bmc3
                _txt(r[26]),  # mission_assured_access
                _txt(r[27]),  # mission_other
                _txt(r[28]),  # mission_not_specified
                _txt(r[29]),  # justification
                _txt(r[30]),  # link_1
                _txt(r[31]),  # link_2
                _txt(r[32]),  # link_3
                _flt(r[33]),  # calculated_lat
                _flt(r[34]),  # calculated_lng
                _flt(r[35]),  # clean_lat
                _flt(r[36]),  # clean_lng
            )
        )
    conn.executemany(
        f"""
        INSERT INTO industry_company (
            priority_country, profile_country, business_unit,
            bu_hq_city, bu_hq_country, founding_year,
            parent_company, parent_hq_city, parent_hq_country,
            primary_value_chain,
            cap_satellite_integration, cap_spacecraft_components,
            cap_payload_sensors, cap_digital_services, cap_launch,
            cap_ground, cap_other,
            primary_mission, sub_mission,
            mission_satcom_pnt, mission_space_sensing,
            mission_sda_combat_power, mission_science_exploration,
            mission_bmc3, mission_assured_access,
            mission_other, mission_not_specified,
            justification, link_1, link_2, link_3,
            calculated_lat, calculated_lng, clean_lat, clean_lng
        ) VALUES ({", ".join(["?"] * 35)})
        """,
        rows,
    )
    return len(rows)


def _load_investment_outlook(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 4):
        country = _txt(r[1])
        mission = _txt(r[2])
        if not country or not mission:
            continue
        rows.append(
            (
                country,
                mission,
                _txt(r[3]),
                _flt(r[4]),
                _txt(r[5]),
                _flt(r[6]),
                _txt(r[7]),
                _txt(r[8]),
                _txt(r[9]),
                _txt(r[10]),
                _txt(r[11]),
                _txt(r[12]),
            )
        )
    conn.executemany(
        """
        INSERT OR REPLACE INTO investment_outlook (
            country, mission, relevant_organizations,
            near_term_score, near_term_priorities,
            long_term_score, long_term_priorities,
            source_1, source_2, source_3,
            ready_for_review, comments
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_launch_sites(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 3):
        # Some Antarctic / extraterritorial sites have no country but are
        # otherwise complete entries — keep them with country=NULL.
        if not any(r[:7]):
            continue
        rows.append(
            (
                _txt(r[0]),
                _txt(r[1]),
                _txt(r[2]),
                _txt(r[3]),
                _txt(r[4]),
                _txt(r[5]),
                _txt(r[6]),
                _flt(r[12]),
                _flt(r[13]),
            )
        )
    conn.executemany(
        """
        INSERT INTO launch_site
            (country, type, status, full_name, infrastructure, name, location, latitude, longitude)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_cities(conn, ws) -> int:
    rows = []
    seen: set[str] = set()
    for r in _rows(ws, 2):
        cid = _txt(r[12])
        if not cid or cid in seen:
            continue
        seen.add(cid)
        rows.append(
            (
                cid,
                _txt(r[0]),
                _txt(r[1]),
                _txt(r[4]),
                _txt(r[7]),
                _txt(r[8]),
                _txt(r[9]),
                _txt(r[10]),
                _int(r[11]),
                _flt(r[2]),
                _flt(r[3]),
            )
        )
    conn.executemany(
        """
        INSERT OR REPLACE INTO city
            (id, city, city_ascii, country, iso2, iso3, admin_name, capital, population, lat, lng)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def _load_vc_deals(conn, ws) -> int:
    """Subset the 166-column PitchBook export down to the analytical core."""
    rows = []
    seen: set[str] = set()
    for r in _rows(ws, 2):
        deal_id = _txt(r[0])
        if not deal_id or deal_id in seen:
            continue
        seen.add(deal_id)
        rows.append(
            (
                deal_id,
                _txt(r[1]),  # company
                _txt(r[2]),  # company_id
                _txt(r[4]),  # description
                _txt(r[6]),  # primary_industry_sector
                _txt(r[7]),  # primary_industry_group
                _txt(r[8]),  # primary_industry_code
                _txt(r[10]),  # verticals
                _txt(r[12]),  # current_financing_status
                _txt(r[13]),  # current_business_status
                _txt(r[23]),  # announced_date
                _txt(r[24]),  # deal_date
                _flt(r[25]),  # deal_size
                _flt(r[27]),  # pre_money_valuation
                _flt(r[28]),  # post_valuation
                _txt(r[35]),  # series
                _txt(r[36]),  # deal_type
                _txt(r[49]),  # deal_status
                _int(r[53]),  # num_investors
            )
        )
    conn.executemany(
        f"""
        INSERT OR REPLACE INTO vc_deal (
            deal_id, company, company_id, description,
            primary_industry_sector, primary_industry_group, primary_industry_code,
            verticals, current_financing_status, current_business_status,
            announced_date, deal_date,
            deal_size_musd, pre_money_valuation_musd, post_valuation_musd,
            series, deal_type, deal_status, num_investors
        ) VALUES ({", ".join(["?"] * 19)})
        """,
        rows,
    )
    return len(rows)


def _load_partnership_sources(conn, ws) -> int:
    rows = []
    for r in _rows(ws, 2):
        src = _txt(r[0])
        cnt = _int(r[1])
        if src and cnt is not None:
            rows.append((src, cnt))
    conn.executemany(
        "INSERT OR REPLACE INTO partnership_source_tally(source, count) VALUES (?, ?)",
        rows,
    )
    return len(rows)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


# Order matters only for FK-dependent loads. With our string-PK references and
# the loader writing all reference data first, sheet order is otherwise free.
SHEET_LOADERS = [
    ("Universal Legend", _load_universal_legend),
    ("Budget Toplines", _load_budget_toplines),
    ("Defense Spending", _load_defense_spending),
    ("Space Program Investments", _load_space_program_investments),
    ("Industrial Base Scores", _load_industrial_base_scores),
    ("Space Assets", _load_space_assets),
    ("International Partnerships", _load_partnerships),
    ("Industry List", _load_industry),
    ("Investment Outlook", _load_investment_outlook),
    ("Launch Sites", _load_launch_sites),
    ("LatLong Auto-Tagging", _load_cities),
    ("VC Investments", _load_vc_deals),
    ("Partnership Sources", _load_partnership_sources),
]


def load_all(xlsx_path: str | Path, db_path: str | Path) -> dict[str, int]:
    """Initialize the schema, load taxonomy + every sheet, and return row counts."""
    counts: dict[str, int] = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    with db.connect(db_path) as conn:
        db.init_schema(conn)
        _load_taxonomy(conn)
        counts["[taxonomy]"] = sum(
            conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            for t in (
                "country",
                "partnership_type",
                "business_model",
                "mission_type",
                "mass_class",
                "partnership_strength_lookup",
            )
        )
        for sheet_name, loader in SHEET_LOADERS:
            ws = wb[sheet_name]
            counts[sheet_name] = loader(conn, ws)
            conn.commit()
    return counts
