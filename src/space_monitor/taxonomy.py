"""Typed access to the controlled vocabularies and scoring rubrics from the
workbook's `Data Validation Lists` sheet.

The JSON file at ``data/taxonomy.json`` is the source of truth at runtime; it is
regenerated from the workbook by :func:`extract_to_json`. Keeping the taxonomy
in JSON (instead of hand-coded enums) means future workbook updates flow
through one extractor instead of code changes scattered across the package.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from functools import lru_cache
from importlib import resources
from pathlib import Path
from typing import Iterable

import openpyxl


# ---------------------------------------------------------------------------
# Typed records
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Country:
    name: str
    cocom: str | None
    priority: bool


@dataclass(frozen=True)
class ScoredTerm:
    name: str
    score: float


@dataclass(frozen=True)
class MassClass:
    name: str
    lower_kg: float
    upper_kg: float


@dataclass(frozen=True)
class StrengthLookup:
    partnership_type: str
    business_model: str
    mission_type: str
    score: int


@dataclass(frozen=True)
class ScoreDefinition:
    score: int
    definition: str


@dataclass(frozen=True)
class Taxonomy:
    cocoms: tuple[str, ...]
    countries: tuple[Country, ...]
    report_types: tuple[str, ...]
    partnership_types: tuple[ScoredTerm, ...]
    levels_of_commitment: tuple[ScoredTerm, ...]
    business_models: tuple[ScoredTerm, ...]
    mission_types: tuple[ScoredTerm, ...]
    relationship_types: tuple[str, ...]
    operator_types: tuple[str, ...]
    organization_types: tuple[str, ...]
    sovereign_statuses: tuple[str, ...]
    mass_classes: tuple[MassClass, ...]
    orbits: tuple[str, ...]
    mission_areas: tuple[str, ...]
    sub_missions: tuple[str, ...]
    value_chain_segments: tuple[str, ...]
    industry_score_definitions: tuple[ScoreDefinition, ...]
    industrial_base_score_definitions: tuple[ScoreDefinition, ...]
    partnership_strength_lookup: tuple[StrengthLookup, ...]

    def country(self, name: str) -> Country | None:
        for c in self.countries:
            if c.name == name:
                return c
        return None

    def score_for_strength(
        self, partnership_type: str, business_model: str, mission_type: str
    ) -> int | None:
        for row in self.partnership_strength_lookup:
            if (
                row.partnership_type == partnership_type
                and row.business_model == business_model
                and row.mission_type == mission_type
            ):
                return row.score
        return None


# ---------------------------------------------------------------------------
# Loader (runtime use)
# ---------------------------------------------------------------------------


def _data_path() -> Path:
    return Path(resources.files("space_monitor") / "data" / "taxonomy.json")


@lru_cache(maxsize=1)
def load() -> Taxonomy:
    """Return the Taxonomy parsed from the bundled JSON file."""
    with _data_path().open() as fh:
        raw = json.load(fh)
    return Taxonomy(
        cocoms=tuple(raw["cocoms"]),
        countries=tuple(Country(**c) for c in raw["countries"]),
        report_types=tuple(raw["report_types"]),
        partnership_types=tuple(ScoredTerm(**t) for t in raw["partnership_types"]),
        levels_of_commitment=tuple(ScoredTerm(**t) for t in raw["levels_of_commitment"]),
        business_models=tuple(ScoredTerm(**t) for t in raw["business_models"]),
        mission_types=tuple(ScoredTerm(**t) for t in raw["mission_types"]),
        relationship_types=tuple(raw["relationship_types"]),
        operator_types=tuple(raw["operator_types"]),
        organization_types=tuple(raw["organization_types"]),
        sovereign_statuses=tuple(raw["sovereign_statuses"]),
        mass_classes=tuple(MassClass(**m) for m in raw["mass_classes"]),
        orbits=tuple(raw["orbits"]),
        mission_areas=tuple(raw["mission_areas"]),
        sub_missions=tuple(raw["sub_missions"]),
        value_chain_segments=tuple(raw["value_chain_segments"]),
        industry_score_definitions=tuple(
            ScoreDefinition(**s) for s in raw["industry_score_definitions"]
        ),
        industrial_base_score_definitions=tuple(
            ScoreDefinition(**s) for s in raw["industrial_base_score_definitions"]
        ),
        partnership_strength_lookup=tuple(
            StrengthLookup(**s) for s in raw["partnership_strength_lookup"]
        ),
    )


# ---------------------------------------------------------------------------
# Extractor (build-time use, run via the CLI)
# ---------------------------------------------------------------------------


def _column(rows: list[tuple], col_index: int) -> list:
    """Return non-empty values from a single column (0-indexed)."""
    out = []
    for r in rows:
        if col_index < len(r):
            v = r[col_index]
            if v not in (None, ""):
                out.append(v)
    return out


def _strip(values: Iterable) -> list:
    return [v.strip() if isinstance(v, str) else v for v in values]


def extract_to_json(xlsx_path: str | Path, out_path: str | Path | None = None) -> Path:
    """Parse the `Data Validation Lists` sheet and write `taxonomy.json`.

    Column indices below match the layout observed in
    ``Space_Dashboard_Hardcopy.xlsx`` (headers in row 3, data starting row 4).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Data Validation Lists"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    cocoms = sorted({c for c in _strip(_column(rows, 0))})
    cocom_set = set(cocoms)

    countries_raw = _strip(_column(rows, 1))
    priority_raw = _strip(_column(rows, 2))
    # COCOM-per-country isn't laid out as a paired column in this sheet; the
    # canonical mapping lives in Universal Legend. Capture priority here, then
    # populate cocom from Universal Legend if available.
    legend = wb["Universal Legend"]
    cocom_by_country = {}
    for r in legend.iter_rows(min_row=2, values_only=True):
        if r and r[0] and r[1]:
            cocom_by_country[str(r[0]).strip()] = str(r[1]).strip()

    countries = []
    for name, prio in zip(countries_raw, priority_raw):
        raw_cocom = cocom_by_country.get(name)
        # Drop sentinel values that aren't real COCOMs.
        cocom = raw_cocom if raw_cocom in cocom_set else None
        countries.append(
            {
                "name": name,
                "cocom": cocom,
                "priority": prio == "Yes",
            }
        )

    report_types = _strip(_column(rows, 3))

    def paired(name_col: int, score_col: int) -> list[dict]:
        names = _strip(_column(rows, name_col))
        scores = _column(rows, score_col)
        return [{"name": n, "score": s} for n, s in zip(names, scores)]

    partnership_types = paired(5, 6)
    levels_of_commitment = paired(7, 8)
    business_models = paired(9, 10)
    mission_types = paired(11, 12)

    # Strength lookup (cols 14, 15, 16 -> 17). Concat in col 18 is the join key.
    strength_rows = []
    for r in rows:
        pt = r[13] if len(r) > 13 else None
        bm = r[14] if len(r) > 14 else None
        mt = r[15] if len(r) > 15 else None
        sc = r[16] if len(r) > 16 else None
        if pt and bm and mt and sc is not None:
            strength_rows.append(
                {
                    "partnership_type": str(pt).strip(),
                    "business_model": str(bm).strip(),
                    "mission_type": str(mt).strip(),
                    "score": int(sc),
                }
            )

    relationship_types = _strip(_column(rows, 19))
    operator_types = _strip(_column(rows, 20))
    organization_types = _strip(_column(rows, 21))
    sovereign_statuses = _strip(_column(rows, 24))

    mass_classes = []
    names = _strip(_column(rows, 25))
    lows = _column(rows, 26)
    highs = _column(rows, 27)
    for n, lo, hi in zip(names, lows, highs):
        mass_classes.append({"name": n, "lower_kg": float(lo), "upper_kg": float(hi)})

    orbits = _strip(_column(rows, 28))
    mission_areas = _strip(_column(rows, 30))
    sub_missions = _strip(_column(rows, 40))
    # Value chain has two adjacent columns (42 + 43); use the longer one.
    vcs = _strip(_column(rows, 42))
    value_chain_segments = list(dict.fromkeys(vcs))  # de-dupe, preserve order

    industry_score_scores = _column(rows, 43)
    industry_score_defs = _strip(_column(rows, 44))
    industry_score_definitions = [
        {"score": int(s), "definition": d}
        for s, d in zip(industry_score_scores, industry_score_defs)
    ]
    industrial_base_defs = _strip(_column(rows, 45))
    industrial_base_score_definitions = [
        {"score": i, "definition": d} for i, d in enumerate(industrial_base_defs)
    ]

    payload = {
        "cocoms": cocoms,
        "countries": countries,
        "report_types": report_types,
        "partnership_types": partnership_types,
        "levels_of_commitment": levels_of_commitment,
        "business_models": business_models,
        "mission_types": mission_types,
        "relationship_types": relationship_types,
        "operator_types": operator_types,
        "organization_types": organization_types,
        "sovereign_statuses": sovereign_statuses,
        "mass_classes": mass_classes,
        "orbits": orbits,
        "mission_areas": mission_areas,
        "sub_missions": sub_missions,
        "value_chain_segments": value_chain_segments,
        "industry_score_definitions": industry_score_definitions,
        "industrial_base_score_definitions": industrial_base_score_definitions,
        "partnership_strength_lookup": strength_rows,
    }

    out = Path(out_path) if out_path else _data_path()
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w") as fh:
        json.dump(payload, fh, indent=2, sort_keys=False)
    load.cache_clear()
    return out
